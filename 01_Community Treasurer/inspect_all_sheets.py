import openpyxl
import sys

def main():
    wb = openpyxl.load_workbook('01_Import/inbox/Lidcombe P&C General Ledger - 2026.xlsx')
    sheets = ['2022-2023', '2023-2024', '2024', '2025', '2026']
    for sheet_name in sheets:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"--- {sheet_name} ---")
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                if i <= 10: # Print first 10 rows
                    print(row)
            print()

if __name__ == "__main__":
    main()
