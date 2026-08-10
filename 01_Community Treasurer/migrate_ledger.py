import openpyxl
from datetime import datetime

def format_date(dt):
    if isinstance(dt, datetime):
        return dt.strftime('%d/%m/%Y')
    return str(dt)

def migrate():
    wb = openpyxl.load_workbook('01_Import/inbox/Lidcombe P&C General Ledger - 2026.xlsx', data_only=True)
    sheets = ['2022-2023', '2023-2024', '2024', '2025', '2026']

    # We will write to two different files
    with open('ledger-0259-migration.md', 'w') as f0259, \
         open('ledger-1844-migration.md', 'w') as f1844:
        
        f0259.write("# Ledger (Account 00900259)\n\n")
        f0259.write("| Date | Description | Category | Amount | Running Balance | Source | Reconciled Period |\n")
        f0259.write("|---|---|---|---|---|---|---|\n")
        
        f1844.write("# Ledger (Account 10741844)\n\n")
        f1844.write("| Date | Description | Category | Amount | Running Balance | Source | Reconciled Period |\n")
        f1844.write("|---|---|---|---|---|---|---|\n")

        for sheet_name in sheets:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            
            # Find header row
            header_row_idx = -1
            headers_found = None
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                if 'Date' in row and 'Description' in row:
                    header_row_idx = i
                    headers_found = list(row)
                    break
            
            if header_row_idx == -1:
                continue

            # Find indices for General Account (0259) - First occurrence
            idx_date_0259 = -1
            idx_desc_0259 = -1
            idx_cat_0259 = -1
            idx_debit_0259 = -1
            idx_credit_0259 = -1
            idx_balance_0259 = -1
            
            # Find indices for Single Sign Account (1844) - Second occurrence
            idx_date_1844 = -1
            idx_desc_1844 = -1
            idx_cat_1844 = -1
            idx_debit_1844 = -1
            idx_credit_1844 = -1
            idx_balance_1844 = -1
            
            date_indices = []
            desc_indices = []
            cat_indices = []
            debit_indices = []
            credit_indices = []
            balance_indices = []

            for i, h in enumerate(headers_found):
                h_low = h.lower() if h else ""
                if 'date' in h_low: date_indices.append(i)
                if 'description' in h_low: desc_indices.append(i)
                if 'category' in h_low: cat_indices.append(i)
                if 'debit' in h_low: debit_indices.append(i)
                if 'credit' in h_low: credit_indices.append(i)
                if 'balance' in h_low: balance_indices.append(i)

            if len(date_indices) >= 1:
                idx_date_0259 = date_indices[0]
                idx_desc_0259 = desc_indices[0] if len(desc_indices) >= 1 else -1
                idx_cat_0259 = cat_indices[0] if len(cat_indices) >= 1 else -1
                idx_debit_0259 = debit_indices[0] if len(debit_indices) >= 1 else -1
                idx_credit_0259 = credit_indices[0] if len(credit_indices) >= 1 else -1
                idx_balance_0259 = balance_indices[0] if len(balance_indices) >= 1 else -1

            if len(date_indices) >= 2:
                idx_date_1844 = date_indices[1]
                idx_desc_1844 = desc_indices[1] if len(desc_indices) >= 2 else -1
                idx_cat_1844 = cat_indices[1] if len(cat_indices) >= 2 else -1
                idx_debit_1844 = debit_indices[1] if len(debit_indices) >= 2 else -1
                idx_credit_1844 = credit_indices[1] if len(credit_indices) >= 2 else -1
                idx_balance_1844 = balance_indices[1] if len(balance_indices) >= 2 else -1

            for row_idx in range(header_row_idx + 1, ws.max_row + 1):
                rows_gen = ws.iter_rows(values_only=True, min_row=row_idx, max_row=row_idx)
                row_values = next(rows_gen)
                
                # Process 0259
                if idx_date_0259 != -1 and row_values[idx_date_0259] is not None:
                    d = format_date(row_values[idx_date_0259])
                    desc = str(row_values[idx_desc_0259]) if idx_desc_0259 != -1 and row_values[idx_desc_0259] is not None else ""
                    cat = str(row_values[idx_cat_0259]) if idx_cat_0259 != -1 and row_values[idx_cat_0259] is not None else ""
                    db = float(row_values[idx_debit_0259]) if idx_debit_0259 != -1 and row_values[idx_debit_0259] is not None else 0.0
                    cr = float(row_values[idx_credit_0259]) if idx_credit_0259 != -1 and row_values[idx_credit_0259] is not None else 0.0
                else:
                    d, desc, cat, db, cr = "", "", "", 0.0, 0.0
                
                bal_0259 = float(row_values[idx_balance_0259]) if idx_balance_0259 != -1 and row_values[idx_balance_0259] is not None else 0.0
                amt_0259 = cr - db
                f0259.write(f"| {d} | {desc} | {cat} | {amt_0259:,.2f} | {bal_0259:,.2f} | Excel Sheet {sheet_name} | Historical |\n")

                # Process 1844
                if idx_date_1844 != -1 and row_values[idx_date_1844] is not None:
                    d = format_date(row_values[idx_date_1844])
                    desc = str(row_values[idx_desc_1844]) if idx_desc_1844 != -1 and row_values[idx_desc_1844] is not None else ""
                    cat = str(row_values[idx_cat_1844]) if idx_cat_1844 != -1 and row_values[idx_cat_1844] is not None else ""
                    db = float(row_values[idx_debit_1844]) if idx_debit_1844 != -1 and row_values[idx_debit_1844] is not None else 0.0
                    cr = float(row_values[idx_credit_1844]) if idx_credit_1844 != -1 and row_values[idx_credit_1844] is not None else 0.0
                else:
                    d, desc, cat, db, cr = "", "", "", 0.0, 0.0
                
                bal_1844 = float(row_values[idx_balance_1844]) if idx_balance_1844 != -1 and row_values[idx_balance_1844] is not None else 0.0
                amt_1844 = cr - db
                f1844.write(f"| {d} | {desc} | {cat} | {amt_1844:,.2f} | {bal_1844:,.2f} | Excel Sheet {sheet_name} | Historical |\n")

if __name__ == "__main__":
    migrate()
