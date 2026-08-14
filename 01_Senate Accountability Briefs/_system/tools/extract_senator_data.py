"""
Stage 1 (extract/fill): pull a senator's data from the source workbook and
contact list, apply the vote-position and portfolio-relevance rules, and
stamp/update records/{slug}/record.md from _templates/senator-record.md.

This does NOT render a docx and does NOT mark anything validated - a human
still has to open the record, check the Portfolio relevance / Vote position
sections, and tick the validation checklist.

Usage (choose exactly one selection mode):
    python extract_senator_data.py "Katherine Gallagher" "Pauline Hanson" ...
    python extract_senator_data.py --party ALP
    python extract_senator_data.py --state QLD
    python extract_senator_data.py --all
"""
import re
import sys
from pathlib import Path

import openpyxl
import yaml
from pypdf import PdfReader

BASE = Path(__file__).resolve().parent.parent.parent  # _system/tools -> _system -> workspace root
REF = BASE / "01_reference"
WORKBOOK = REF / "source-data" / "Senate ndis table completed.xlsx"
CONTACTS = REF / "source-data" / "senate contact list.pdf"
VOTE_RULES = REF / "vote-position-rules.md"
RECORD_TEMPLATE = BASE / "_templates" / "senator-record.md"
RECORDS_DIR = BASE / "records"

STATE_NAMES = {
    "ACT": "Australian Capital Territory", "NSW": "New South Wales", "NT": "Northern Territory",
    "QLD": "Queensland", "SA": "South Australia", "TAS": "Tasmania", "VIC": "Victoria", "WA": "Western Australia",
}
PARTY_FULL = {
    "ALP": "Australian Labor Party", "LP": "Liberal Party", "NATS": "The Nationals", "NP": "The Nationals",
    "LNP": "Liberal National Party", "AG": "Australian Greens", "ON": "Pauline Hanson's One Nation",
    "UAP": "United Australia Party", "IND": "Independent", "JLN": "Jacqui Lambie Network",
    "CLP": "Country Liberal Party", "AV": "Australia's Voice",
}
GOVT_BLOC_ABBR = {"ALP", "LP", "NATS", "NP", "LNP", "ON", "CLP"}
GREENS_ABBR = {"AG"}

# Kept in sync with 01_reference/portfolio-relevance.md's category table - update both if categories change.
RELEVANT_PORTFOLIO_KEYWORDS = [
    "Women", "NDIS", "Disability", "Health", "Ageing", "Aged Care",
    "Social Services", "Government Services", "Finance", "Treasury",
    "Indigenous", "Regional Australia", "Regional Health",
]
PROCEDURAL_ONLY_MARKERS = ["Whip", "Manager of", "Deputy President", "President)", "Leader of the", "Deputy Leader", "Chair of"]

STATE_ABBR_RE = r"(?:NSW|VIC|QLD|WA|SA|TAS|NT|ACT)"
STATE_PARTY_STOP = STATE_ABBR_RE + r"\s+[A-Z]{2,5}\b"


def normalize_name(name: str) -> str:
    name = re.sub(r"^(Senator\s+)?(the\s+Hon\s+)?", "", name, flags=re.I)
    name = re.sub(r"\s*\([^)]*\)", "", name)
    words = [w for w in re.sub(r"\s+", " ", name).strip().split(" ") if w]
    if len(words) < 2:
        return name.strip().lower()
    return f"{words[0]} {words[-1]}".lower()


def slugify(first_last_key: str) -> str:
    return first_last_key.replace(" ", "-")


def load_workbook_rows():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        rows[normalize_name(rec["Senator"])] = rec
    return rows


def load_contacts():
    reader = PdfReader(str(CONTACTS))
    full_text = "\n".join(p.extract_text() for p in reader.pages)
    marker_re = re.compile(r"\n\**\d+\s+[A-Za-z'\u2019\-]+(?:\s+[A-Za-z'\u2019\-]+)?,\s+Senator\b")
    markers = list(marker_re.finditer(full_text))

    entries = {}
    for i, m in enumerate(markers):
        block_start = m.start() + 1
        block_end = markers[i + 1].start() if i + 1 < len(markers) else len(full_text)
        block = full_text[block_start:block_end]

        header_match = re.match(
            r"\**\d+\s+([A-Za-z'\u2019\-]+(?:\s+[A-Za-z'\u2019\-]+)?),\s+Senator\s+(.+?)\s*(?=" + STATE_PARTY_STOP + ")",
            block, re.S,
        )
        if not header_match:
            print(f"WARN: could not find name/state in block: {block[:80]!r}")
            continue
        last, name_and_title = header_match.groups()
        lines = [l.strip() for l in name_and_title.split("\n") if l.strip()]
        first_full_raw = lines[0] if lines else ""
        # Drop lone-nickname lines like "(Katy)" / "(Tim)" - not a role/title.
        title_lines = [l for l in lines[1:] if not re.match(r"^\([A-Za-z]+\)$", l)]
        first_full = re.sub(r"\s*\([^)]*\)", "", first_full_raw).strip()
        full_name = f"{first_full} {last}".strip()

        state_party = re.search(rf"\b({STATE_ABBR_RE})\s+([A-Z]{{2,5}})\b", block)
        email = re.search(r"Email:\s*([\w.\-\u2019']+@aph\.gov\.au)", block)
        if not state_party or not email:
            print(f"WARN: could not parse contact block for {full_name}")
            continue

        entries[normalize_name(full_name)] = {
            "state": state_party.group(1),
            "party": state_party.group(2),
            "email": email.group(1),
            "title": "; ".join(title_lines),
        }
    return entries


def load_vote_overrides():
    text = VOTE_RULES.read_text(encoding="utf-8")
    section = text.split("## Known overrides")[1].split("## Fixed immediate ask")[0]
    overrides = {}
    for line in section.splitlines():
        m = re.match(r"\|\s*(.+?)\s*\|\s*(.+?)\s*\|\s*(.+?)\s*\|$", line)
        if not m or m.group(1) in ("Senator", "---"):
            continue
        name, position, source = m.groups()
        overrides[normalize_name(name)] = (position, source)
    return overrides


def vote_position(senator_key, party_abbr, overrides):
    if senator_key in overrides:
        position, source = overrides[senator_key]
        return f"{position} ({source})", "override"
    if party_abbr in GOVT_BLOC_ABBR:
        return f"Expected {PARTY_FULL.get(party_abbr, party_abbr)} support, subject to confirmation", "rule"
    if party_abbr in GREENS_ABBR:
        return "Expected opposition", "rule"
    return "Unknown", "rule"


def portfolio_relevance(title: str):
    """Returns (relevant, matched_line_or_None, note). `title` may hold several
    ';'-joined role/title lines - relevance is checked per line so the specific
    matching line (not the whole title dump) becomes the brief's portfolio field."""
    if not title:
        return False, None, "No portfolio held (per source-data snapshot)"
    lines = [l.strip() for l in title.split(";") if l.strip()]
    for line in lines:
        if any(kw.lower() in line.lower() for kw in RELEVANT_PORTFOLIO_KEYWORDS):
            return True, line, "Matches a relevant category in portfolio-relevance.md - CONFIRM the specific mechanism by hand, don't ship this line as-is"
    is_procedural = any(pm.lower() in title.lower() for pm in PROCEDURAL_ONLY_MARKERS)
    note = "Procedural Senate role, not a policy portfolio" if is_procedural else "No relevant category matched"
    return False, None, note


def fmt_money_billion(v):
    return f"{v / 1_000_000_000:.3f}"


def fmt_money_million(v):
    return f"{v / 1_000_000:.1f}"


def build_record(senator_key, wb_row, contact, overrides):
    senator_display = wb_row["Senator"].replace("Senator the Hon ", "").replace("Senator ", "")
    state_abbr = wb_row["State"]
    state_full = STATE_NAMES.get(state_abbr, state_abbr)
    participants = int(wb_row["Active Participants"])
    providers = int(wb_row["Active Providers"])
    total_electors = int(wb_row["State Electors"])
    female_electors = int(wb_row["State Female Electors"])
    male_electors = int(wb_row["State Male Electors"])
    children = int(wb_row["0-14 Years"])
    autism = int(wb_row["Autism"])
    annual_funding = wb_row["Estimated Annual Funding Flow"]

    carers = participants * 2
    family = participants * 4
    community = participants * 6
    total_footprint = participants + carers + family + providers + community

    position, source = vote_position(senator_key, contact["party"], overrides)
    relevant, matched_portfolio, relevance_note = portfolio_relevance(contact["title"])
    full_title_raw = contact["title"] or "—"

    slug = slugify(senator_key)

    fm = {
        "type": "senator-record",
        "slug": slug,
        "senator_name": senator_display,
        "party": PARTY_FULL.get(contact["party"], contact["party"]),
        "party_abbr": contact["party"],
        "state": state_full,
        "state_abbr": state_abbr,
        "email": contact["email"],
        "portfolio": matched_portfolio,
        "vote_position": position,
        "vote_position_source": source,
        "status": "computed",
        "validated": False,
        "custom-width": 100,
    }

    body = f"""
# {senator_display} — record

Field meanings and sources: [../01_reference/brief-template.md](../01_reference/brief-template.md)

## Pulled data
*(from `01_reference/source-data/` — regenerate via `_system/tools/extract_senator_data.py`, don't hand-edit these)*

- Total Electors: {total_electors:,}
- Female Electors: {female_electors:,}
- Male Electors: {male_electors:,}
- Active NDIS Participants: {participants:,}
- Children (0-14 Years): {children:,}
- Autism Participants: {autism:,}
- Psychosocial Participants: {int(wb_row['Psychosocial Disability']):,}
- First Nations Participants: {int(wb_row['First Nations']):,}
- CALD Participants: {int(wb_row['CALD']):,}
- Active Providers: {providers:,}
- Annual NDIS Funding Exposure ($ billion): {fmt_money_billion(annual_funding)}

## Computed fields
*(formulas: [../01_reference/brief-template.md](../01_reference/brief-template.md) field legend)*

- Female Share (%): {female_electors / total_electors * 100:.1f}
- Child Share (%): {children / participants * 100:.1f}
- Autism Share (%): {autism / participants * 100:.1f}
- 1% Reduction ($ million): {fmt_money_million(annual_funding * 0.01)}
- 5% Reduction ($ million): {fmt_money_million(annual_funding * 0.05)}
- 10% Reduction ($ million): {fmt_money_million(annual_funding * 0.10)}
- Carers Estimate: {carers:,}
- Family Estimate: {family:,}
- Community Estimate: {community:,}
- Total Footprint Estimate: {total_footprint:,}

## Portfolio relevance check
*(rule: [../01_reference/portfolio-relevance.md](../01_reference/portfolio-relevance.md); snapshot lookup: [../01_reference/senator-portfolios.md](../01_reference/senator-portfolios.md) — confirm currency, don't trust the snapshot blindly)*

- Full title on public record (context only, not what goes in the brief): {full_title_raw}
- Matched portfolio (used as `[PORTFOLIO/MINISTERIAL ROLE]` if relevant): {matched_portfolio or "—"}
- Relevant to this bill? {"Y" if relevant else "N"} — {relevance_note}
- If yes, the specific mechanism (not a generic paragraph): {"TODO - human to write, mirroring how the Women lens ties SCCP cuts to unpaid care" if relevant else "n/a"}

## Vote position
*(rule: [../01_reference/vote-position-rules.md](../01_reference/vote-position-rules.md))*

- Position: {position}
- Source: {source}

## Validation checklist

- [ ] Elector data matches `01_reference/source-data/Senate ndis table completed.xlsx`
- [ ] NDIS participant data matches the workbook
- [ ] Party/email match `01_reference/source-data/senate contact list.pdf`
- [ ] Portfolio relevance actively checked (not defaulted to blank)
- [ ] Vote position checked against the rule + any known override
- [ ] Immediate ask correct (fixed text, unless a reason to override)
- [ ] No leftover `[BRACKET]` placeholders

## Human notes

"""
    return slug, fm, body


def write_record(slug, frontmatter, body):
    record_dir = RECORDS_DIR / slug
    record_dir.mkdir(parents=True, exist_ok=True)
    out_path = record_dir / "record.md"
    fm_text = yaml.dump(frontmatter, sort_keys=False, allow_unicode=True)
    out_path.write_text(f"---\n{fm_text}---\n{body}", encoding="utf-8")
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    sys.path.insert(0, str(Path(__file__).parent))
    from _selection import parse_selection_args

    wb_rows = load_workbook_rows()
    contacts = load_contacts()
    overrides = load_vote_overrides()
    print(f"Workbook senators: {len(wb_rows)}  Contact list senators: {len(contacts)}  Vote overrides: {len(overrides)}")

    default_targets = ["katherine gallagher", "pauline hanson", "david pocock", "jordon steele-john"]
    selection, defaults = parse_selection_args(sys.argv[1:], default_names=default_targets)
    selection.names = [normalize_name(n) for n in selection.names]
    available = {k: {"party": v["party"], "state": v["state"]} for k, v in contacts.items()}
    targets = selection.resolve(available, default_keys=defaults)
    print(f"Selected {len(targets)} senator(s)")

    for key in targets:
        if key not in wb_rows:
            print(f"SKIP (not in workbook): {key}")
            continue
        if key not in contacts:
            print(f"SKIP (not in contact list): {key}")
            continue
        slug, fm, body = build_record(key, wb_rows[key], contacts[key], overrides)
        write_record(slug, fm, body)
